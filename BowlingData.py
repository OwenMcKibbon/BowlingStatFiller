
from urllib.request import urlopen
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill
from pandas import DataFrame
import numpy as np
import os

from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename

def backToExcel(dataFrames: dict) -> None:
    save_path = asksaveasfilename(title="Save As", defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    file_exists = os.path.isfile(save_path)
    if not save_path:
        print("No file selected")
        return
    else:
        with pd.ExcelWriter(save_path, engine="openpyxl", mode='a' if file_exists else 'w') as writer:
            for sheet, df in dataFrames.items():
                try:
                    formattedDf = df.replace("NEW_X", "X")
                    formattedDf.to_excel(writer, sheet_name=sheet, index=False, startrow=0, startcol=0)
                    ws = writer.book[sheet]

                    # Apply conditional formatting to color "X" cells yellow
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row, start=1):  # Excel columns start at 1
                            if str(value).strip().upper() == "NEW_X":
                                cell = ws.cell(row=row_idx + 2, column=col_idx)  # +2 for header and zero-indexing
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    print("Excel sheet " + sheet + " created")
                except Exception as e:
                    print("Error creating workbook.")
            """
            awardSummary = []
            for division, bowlerList in awards.items():
                awardSummary.append([division])
                awardSummary.append(["Bowler Name", "New Awards"])
                seenBowlers = set()
                for bowler, awardList in bowlerList:
                    if (division, bowler) not in seenBowlers:
                        seenBowlers.add((division, bowler))
                        uniqueAwards = ", ".join(set(awardList))
                        awardSummary.append([bowler, uniqueAwards])
                awardSummary.append([])

            awardSummaryDf = pd.DataFrame(awardSummary)
            awardSummaryDf.to_excel(writer, sheet_name="Award Summary", index=False, startrow=0, startcol=0)
            print("Excel sheet: Awards Summary created")
            """

        print("Workbook saved to: " + save_path)

class Bowler:

    gender: str = ""
    name: str = ""

    average: float = 0

    highSingle: float = 0
    highSeries: float = 0

    attendance: int = 0

    division: str = ""
    link: str = ""

    stats: list = []
    generalStats: list = []

    def __init__(self, division: str, name: str, link: str):

        self.division = division
        self.name = name
        self.fixName()
        self.link = link
        self.clearStats()
        self.fillGeneralStats()
        self.fillStats()
        self.fillGeneralStats()
        self.setBasicStats()

    def fixName(self):
        if self.name == "JEAN-SEBASIEN GORLEY":
            self.name = "JEAN-SEBASTIEN GORLEY"

    def getName(self):
        return self.name

    def getStats(self):
        return self.stats

    def clearStats(self):
        self.stats = []

    def clearGeneralStats(self):
        self.generalStats = []

    def getGeneralStats(self):
        return self.generalStats

    def setAverage(self):
        avg = 0.0
        for row in self.generalStats:
            if str(row[0]) == "Scratch Average":
                avg = float((row[1]))
                break
        self.average = avg

    def getAverage(self):

        return self.average

    def getHighSeries(self):

        return self.highSeries

    def getHighSingle(self):

        return self.highSingle

    def getGender(self):
        return self.gender

    def getDivision(self):

        return self.division

    def setDivision(self, division):

        self.division = division

    def getAttendance(self):
        return self.attendance


    def fillStats(self):
        page = urlopen(self.link)
        html = page.read().decode("utf-8")
        soup = BeautifulSoup(html, "html.parser")
        tables = soup.find_all("table")
        table = tables[1]

        headers = [header.text for header in table.find_all("th")]

        bowlerStats = [headers]
        if table:
            rows = table.find_all("tr")[1:]
            for row in rows:
                cells = row.find_all("td")
                if cells:
                    bowlerStats.append([cell.text for cell in cells])

        self.stats = bowlerStats


    def fillGeneralStats(self):

        page = urlopen(self.link)
        html = page.read().decode("utf-8")
        soup = BeautifulSoup(html, "html.parser")
        tables = soup.find_all("table")
        table = tables[0]

        if table:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all("td")
                if cells:
                    self.generalStats.append([cell.text for cell in cells])

        #self.generalStats = bowlerStats

    def setBasicStats(self):

        for row in self.generalStats:

            if str(row[0]) == "Gender":
                self.gender = str(row[1])

            if str(row[0]) == "Scratch Average":
                self.average = float(row[1])

            if len(row) > 2 and str(row[2]) == "High Game Scratch (with Info.)":
                self.highSingle = float((row[3])[:3])

            if len(row) > 2 and str(row[2]) == "High Series":
                self.highSeries = float((row[3]))

    def getPOA(self):

        hasPOA = False

        for stat in self.stats[1:]:
            try:
                game1 = int((stat[3])[:3])
                game2 = int((stat[4])[:3])
                game3 = int((stat[5])[:3])
                if (stat[9])[:5] == "-":
                    average = float((stat[6])[:5])
                else:
                    average = float((stat[9])[:6])

                if game1 >= average + 100 or game2 >= average + 100 or game3 >= average + 100:
                    hasPOA = True
                    break
            except ValueError:
                pass
        return hasPOA


    def hasGame(self, lower: int, upper: int):
        for stat in self.stats[1:]:
            try:
                if (lower <= int((stat[3])[:3]) < upper) or (lower <= int((stat[4])[:3]) < upper) or (
                        lower <= int((stat[5])[:3]) < upper):
                    return True
            except ValueError:
                pass
        return False

    def hasTotal(self, lower: int, upper: int):
        for stat in self.stats[1:]:
            if lower <= int(stat[-2]) < upper:
                return True
        return False



class Division:

    name: str = ""

    bowlers: dict = {}
    mensBowlers: dict = {}
    womensBowlers: dict = {}

    topPerformers: dict = {}
    topPerformersMen: dict = {}
    topPerformersWomen: dict = {}

    awardSummary: list = []

    def __init__(self, name: str):
        self.name = name
        if not name == "Special Awards":
            self.setBowlers()
            self.setTopPerformersMen()
            self.setTopPerformersWomen()
            self.setTopPerformers()

    def getDivision(self):
        return self

    def getName(self):
        return self.name

    def addBowler(self, bowler: Bowler):
        self.bowlers.update({bowler.name: bowler})

    def removeBowler(self, bowler: Bowler):
        self.bowlers.pop(bowler.name, None)

    def getBowlers(self):
        return self.bowlers

    def setTopPerformers(self):
        highAverage = 0
        highestAvgBowler = None

        highSeries = 0
        highestSeriesBowler = None

        highSingle = 0
        highestSingleBowler = None
        for bowler in self.bowlers.values():
            if bowler.getAverage() > highAverage:
                highAverage = bowler.getAverage()
                highestAvgBowler = bowler
            if bowler.getHighSeries() > highSeries:
                highSeries = bowler.getHighSeries()
                highestSeriesBowler = bowler
            if bowler.getHighSingle() > highSingle:
                highSingle = bowler.getHighSingle()
                highestSingleBowler = bowler

        self.topPerformers["High Avg"] = highestAvgBowler
        self.topPerformers["High Series"] = highestSeriesBowler
        self.topPerformers["High Single"] = highestSingleBowler

    def setTopPerformersWomen(self):
        highAverage = 0
        highestAvgBowler = None
        highSeries = 0
        highestSeriesBowler = None
        highSingle = 0
        highestSingleBowler = None

        for bowler in self.womensBowlers.values():
            if bowler.getAverage() > highAverage:
                highAverage = bowler.getAverage()
                highestAvgBowler = bowler
            if bowler.getHighSeries() > highSeries:
                highSeries = bowler.getHighSeries()
                highestSeriesBowler = bowler
            if bowler.getHighSingle() > highSingle:
                highSingle = bowler.getHighSingle()
                highestSingleBowler = bowler

        self.topPerformersWomen["High Avg"] = highestAvgBowler
        self.topPerformersWomen["High Series"] = highestSeriesBowler
        self.topPerformersWomen["High Single"] = highestSingleBowler

    def setTopPerformersMen(self):
        highAverage = 0
        highestAvgBowler = None
        highSeries = 0
        highestSeriesBowler = None
        highSingle = 0
        highestSingleBowler = None

        for bowler in self.mensBowlers.values():
            if bowler.getAverage() > highAverage:
                highAverage = bowler.getAverage()
                highestAvgBowler = bowler
            if bowler.getHighSeries() > highSeries:
                highSeries = bowler.getHighSeries()
                highestSeriesBowler = bowler
            if bowler.getHighSingle() > highSingle:
                highSingle = bowler.getHighSingle()
                highestSingleBowler = bowler

        self.topPerformersMen["High Avg"] = highestAvgBowler
        self.topPerformersMen["High Series"] = highestSeriesBowler
        self.topPerformersMen["High Single"] = highestSingleBowler

    def getHighestAvg(self):
        return self.topPerformers["High Avg"]

    def getHighestSeries(self):
        return self.topPerformers["High Series"]

    def getHighestSingle(self):
        return self.topPerformers["High Single"]

    def setBowlers(self):
        division = self.name
        if self.name == 'Senior':
            division = 'Seniors'

        self.bowlers = {}
        self.mensBowlers = {}
        self.womensBowlers = {}

        URL = f"https://bowling.lexerbowling.com/orleansbowlingcentre/ybc{division.lower()}2024-2025/playerlist.htm"

        page = urlopen(URL)
        html = page.read().decode("utf-8")
        soup = BeautifulSoup(html, "html.parser")
        tables = soup.find_all("table")
        mensTable = tables[0]
        womensTable = tables[1]

        self.getTableBowlers(mensTable, division, "Men")
        self.getTableBowlers(womensTable, division, "Women")

    def getTableBowlers(self, table: DataFrame, division: str, gender: str):

        for row in table.find_all("tr")[1:]:
            cells = row.find_all("td")
            if cells:
                link = cells[1].find("a")
                if link:
                    reference = link.get("href")
                    reference = reference[1:]
                    bowlerLink = f"https://bowling.lexerbowling.com/orleansbowlingcentre/ybc{division.lower()}2024-2025" + reference
                    bowler: Bowler = Bowler(self.name, str(cells[1].text), bowlerLink)
                    self.bowlers.update({bowler.name: bowler})
                    if gender == "Men":
                        self.mensBowlers.update({bowler.name: bowler})
                    elif gender == "Women":
                        self.womensBowlers.update({bowler.name: bowler})

    def getWomenTopPerformers(self):
        return self.topPerformersWomen

    def getMenTopPerformers(self):
        return self.topPerformersMen

    def clearTopPerformers(self):
        self.topPerformersMen = {}
        self.topPerformersWomen = {}
        self.topPerformers = {}


class Sheet:

    __wholeFrame: DataFrame
    __division: Division
    __subBantam: Division
    __subJunior: Division
    __subSenior: Division

    __updatedCells: list = []
    __topCells: list = []
    __bottomCells: list = []

    def __init__(self, wholeFrame: DataFrame, division: Division, subDiv1: Division, subDiv2: Division, subDiv3: Division):
        self.__wholeFrame = wholeFrame.copy()

        self.__division = division
        self.__subBantam = subDiv1
        self.__subJunior = subDiv2
        self.__subSenior = subDiv3

        self.__wholeFrame.drop(self.__wholeFrame.columns[-1], axis=1)
        pd.set_option('future.no_silent_downcasting', True)
        self.__wholeFrame.replace(r'^\s*$', np.nan, regex=True, inplace=True)
        self.__wholeFrame.dropna(axis=1, how="all")
        self.fillSheet()

    def fillSheet(self):
        pd.set_option('future.no_silent_downcasting', True)
        self.__wholeFrame.replace(r'^\s*$', np.nan, regex=True, inplace=True)
        self.__wholeFrame.dropna(axis=1, how="all", inplace=True)
        self.__wholeFrame = self.__wholeFrame.fillna('')

        if self.__division.getName() in ["Bantam", "Junior", "Senior"]:
            gameDf, seriesDf = self.splitFrame()
            gameDf = self.appendXs(gameDf, "top")
            seriesDf = self.appendXs(seriesDf, "bottom")
            self.__wholeFrame = self.rejoin(gameDf, seriesDf)

        elif self.__division.getName() == "Special Awards":
            #self.__wholeFrame.replace({'X': '', 'x' : ''}, inplace=True)
            self.__wholeFrame.iloc[:, [i for i in range(self.__wholeFrame.shape[1]) if i != 2]] = \
                self.__wholeFrame.iloc[:, [i for i in range(self.__wholeFrame.shape[1]) if i != 2]].replace(
                    {'X': '', 'x': ''}, inplace=False)

            bantamAwardsDf, juniorAwardsDf, seniorAwardsDf = self.splitSpecial()

            bantamDivision: Division = self.__subBantam
            bantamAwardsDf = self.appendAwardXs(bantamAwardsDf, bantamDivision)

            juniorDivision: Division = self.__subJunior
            juniorAwardsDf = self.appendAwardXs(juniorAwardsDf, juniorDivision)

            seniorDivision: Division = self.__subSenior
            seniorAwardsDf = self.appendAwardXs(seniorAwardsDf, seniorDivision)

            awards = self.rejoin(bantamAwardsDf, juniorAwardsDf)
            finalAwards = self.rejoin(awards, seniorAwardsDf)

            self.__wholeFrame = finalAwards


    def getDivision(self):
        return self.__division

    def splitFrame(self):

        df = self.__wholeFrame
        division = self.__division.getName()

        df.iloc[:, 0] = df.iloc[:, 0].str.strip()
        division = division.strip()

        #secondHeaderRow = df[df.iloc[:, 0] == division].index[0]
        secondHeaderRow = df[df.iloc[:, 0] == division].index[0]

        gameDf = df.iloc[:secondHeaderRow]
        seriesDf = df.iloc[secondHeaderRow + 1:]

        gameDf.columns = df.columns
        seriesDf.columns = df.iloc[secondHeaderRow].values

        seriesDf = seriesDf.reset_index(drop=True)

        return gameDf, seriesDf

    def splitSpecial(self):
        df = self.__wholeFrame
        df.iloc[:, 0] = df.iloc[:, 0].fillna('').str.strip()

        secondHeaderRow = df[df.iloc[:, 0] == "Junior"].index[0]
        thirdHeaderRow = df[df.iloc[:, 0] == "Senior"].index[0]

        bantamSpecialDf = df.iloc[:secondHeaderRow].copy()
        juniorSpecialDf = df.iloc[secondHeaderRow + 1:thirdHeaderRow].copy()
        seniorSpecialDf = df.iloc[thirdHeaderRow + 1:].copy()

        bantamSpecialDf.columns = df.columns

        juniorSpecialDf.columns = ["Junior"] + list(df.columns[1:])

        juniorSpecialDf = juniorSpecialDf.reset_index(drop=True)

        seniorSpecialDf.columns = ["Senior"] + list(df.columns[1:])

        seniorSpecialDf = seniorSpecialDf.reset_index(drop=True)

        return bantamSpecialDf, juniorSpecialDf, seniorSpecialDf

    def rejoin(self, gameDf, seriesDf):

        seriesHeader = seriesDf.columns.to_list()
        seriesRows = seriesDf.values.tolist()

        seriesRows.insert(0, seriesHeader)

        wholeTable = gameDf.values.tolist() + seriesRows

        wholeTableDf = pd.DataFrame(wholeTable, columns=gameDf.columns)

        return wholeTableDf

    def setFrame(self, finalFrame: DataFrame):
        self.__wholeFrame = finalFrame


    def appendAwardXs(self, df: DataFrame, division: Division):

        bowlers = division.getBowlers()

        for index, row in df.iterrows():

            name = str(row.iloc[0]) + " " + str(row.iloc[1])

            bowler = bowlers.get(name)
            if not bowler:
                continue

            for column in range(2, 5): #len(df.columns)
                columnValue = row.iloc[column]
                isX = (columnValue == "X")
                isx = (columnValue == "x")
                if columnValue == "NEW_X":
                    pass

                if (not isX) and (not isx):
                    columnType = df.columns[column]
                    check = False

                    #Checking if POA for currentBowler is >= 100 for most recent week
                    if str(columnType) == "100 Pins over Average":
                        check = bowler.getPOA()

                    #Checking if attendance for currentBowler is equal to every game
                    elif str(columnType) == "Perfect Attendance":
                        check = bowler.getAttendance() == 32

                    elif str(columnType) == "Most Improved":
                        check = False

                    if check:
                        df.iat[index, column] = "NEW_X"


                    else:
                        pass

        women = division.getWomenTopPerformers()

        for key in women.keys():
            value = women[key]
            firstName, lastName = value.getName().split(" ")
            playerRow = df[(df.iloc[:, 0] == firstName) & (df.iloc[:, 1] == lastName)]
            if not playerRow.empty:
                playerIndex = playerRow.index[0]
                if key in df.columns:
                    df.at[playerIndex, key] = 'NEW_X'

        men = division.getMenTopPerformers()

        for key in men.keys():
            value = men[key]
            firstName, lastName = value.getName().split(" ")
            mensRow = df[(df.iloc[:, 0] == firstName) & (df.iloc[:, 1] == lastName)]
            if not mensRow.empty:
                playerIndex = mensRow.index[0]
                if key in df.columns:
                    df.at[playerIndex, key] = 'NEW_X'
        df = df.fillna('')


        return df

    def appendXs(self, df: DataFrame, order: str):

        bowlers = self.__division.getBowlers()

        for index, row in df.iterrows():

            name = str(row.iloc[0]) + " " + str(row.iloc[1])

            bowler = bowlers.get(name)
            if not bowler:
                continue

            #bowlerStats = bowler.getStats()

            #This is for the top half of the dataFrame which has ranges of games instead of totals
            if order == "top":
                for column in range(2, len(df.columns)):
                    columnValue = row.iloc[column]
                    columnValue = columnValue.strip()
                    isX = (columnValue == "X")
                    isx = (columnValue == "x")
                    if columnValue == "NEW_X":
                        pass

                    if (not isX) and (not isx):
                        num = df.columns[column]

                        if num == '':
                            continue
                        #if num.isdigit():
                        try:
                            gameRange = int(num)
                        except ValueError:
                            continue
                        #else:
                            #continue

                        #In the special case that the column header is 400 the possible range is 400-450
                        #Instead of for example 250-275
                        isLast = (gameRange == 400)
                        if isLast:
                            hasGame = bowler.hasGame(gameRange, gameRange + 50)
                        else:
                            hasGame = bowler.hasGame(gameRange, gameRange + 25)

                        #If the bowler has a game in that range then append X
                        if hasGame:
                            df.iat[index, column] = "NEW_X"

            #This is for the bottom half of the data frame which has ranges of totals instead of games
            elif order == "bottom":
                for column in range(2, len(df.columns)):
                    columnValue = row.iloc[column]
                    isX = (columnValue == "X")
                    isx = (columnValue == "x")
                    if columnValue == "NEW_X":
                        pass

                    if (not isX) and (not isx):
                        num = df.columns[column]
                        if num == '':
                            continue

                        try:
                            gameRange = int(num)
                        except ValueError:
                            continue

                        hasTotal = bowler.hasTotal(gameRange, gameRange + 50)
                        if hasTotal:
                            df.iat[index, column] = "NEW_X"

        return df

    def getFrame(self):
        return self.__wholeFrame



def main() -> None:

    root = Tk()
    root.withdraw()

    file_path = askopenfilename(title= "Select the Excel file", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")) )

    if not file_path:
        print("No file selected")
    else:
        try:

            bantamDf = pd.DataFrame(pd.read_excel(file_path, sheet_name=0, engine="openpyxl"))
            bantamDivision: Division = Division("Bantam")
            bantamSheet: Sheet = Sheet(bantamDf, bantamDivision, None, None, None)
            bantamDf = bantamSheet.getFrame()

            juniorDf = pd.DataFrame(pd.read_excel(file_path, sheet_name=1, engine="openpyxl"))
            juniorDivision: Division = Division("Junior")
            juniorSheet: Sheet = Sheet(juniorDf, juniorDivision, None, None, None)
            juniorDf = juniorSheet.getFrame()

            seniorDf = pd.DataFrame(pd.read_excel(file_path, sheet_name=2, engine="openpyxl"))
            seniorDivision: Division = Division("Senior")
            seniorSheet: Sheet = Sheet(seniorDf, seniorDivision, None, None, None)
            seniorDf = seniorSheet.getFrame()

            awardsDf = pd.DataFrame(pd.read_excel(file_path, sheet_name=3, engine="openpyxl"))
            awardsDivision: Division = Division("Special Awards")

            bantamDivision.clearTopPerformers()
            bantamDivision.setTopPerformersMen()
            bantamDivision.setTopPerformersWomen()

            juniorDivision.clearTopPerformers()
            juniorDivision.setTopPerformersMen()
            juniorDivision.setTopPerformersWomen()

            seniorDivision.clearTopPerformers()
            seniorDivision.setTopPerformersMen()
            seniorDivision.setTopPerformersWomen()

            awardSheet: Sheet = Sheet(awardsDf, awardsDivision, bantamDivision, juniorDivision, seniorDivision)
            awardsDf = awardSheet.getFrame()

            sheets = {"Badges Bantam": bantamDf, "Badges Junior": juniorDf, "Badges Senior": seniorDf,
                      "Special Awards": awardsDf}

            backToExcel(sheets)

        except FileNotFoundError:
            print("File could not be located")
        except Exception as e:
            print(e)


if __name__=="__main__":
    main()
